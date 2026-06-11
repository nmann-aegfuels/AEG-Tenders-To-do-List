"""
AEG Fuels — Global Tender Schedule Processor
Azure Function: pulls Global_Tender_Schedule.xlsm from SharePoint via
Microsoft Graph API on a timer → parses all three regional sheets
(EMEA Query, US Query, APAC Query) → writes tenders.json to Blob Storage.

Requirements (requirements.txt):
    azure-functions
    azure-storage-blob
    azure-identity
    openpyxl
    requests

App Settings (Azure portal + local.settings.json):
    AZURE_TENANT_ID          your Microsoft 365 tenant ID
    AZURE_CLIENT_ID          App Registration client ID
    AZURE_CLIENT_SECRET      App Registration client secret
    SHAREPOINT_SITE_ID       SharePoint site ID
    SHAREPOINT_FILE_PATH     e.g. /Shared Documents/Global_Tender_Schedule.xlsm
    AzureWebJobsStorage      Blob Storage connection string
    OUTPUT_CONTAINER         default: tenders-output
    OUTPUT_BLOB_NAME         default: tenders.json
"""

import json
import logging
import os
from datetime import date, datetime, timedelta, timezone
from io import BytesIO

import azure.functions as func
import openpyxl
import requests
from azure.identity import ClientSecretCredential
from azure.storage.blob import BlobServiceClient, ContentSettings

# ── Config ────────────────────────────────────────────────────────────────────
TENANT_ID            = os.environ["AZURE_TENANT_ID"]
CLIENT_ID            = os.environ["AZURE_CLIENT_ID"]
CLIENT_SECRET        = os.environ["AZURE_CLIENT_SECRET"]
SHAREPOINT_SITE_ID   = os.environ["SHAREPOINT_SITE_ID"]
SHAREPOINT_FILE_PATH = os.environ.get("SHAREPOINT_FILE_PATH",
                                      "/Shared Documents/Global_Tender_Schedule.xlsm")
STORAGE_CONNECTION   = os.environ["AzureWebJobsStorage"]
OUTPUT_CONTAINER     = os.environ.get("OUTPUT_CONTAINER", "tenders-output")
OUTPUT_BLOB_NAME     = os.environ.get("OUTPUT_BLOB_NAME", "tenders.json")

GRAPH_SCOPE = "https://graph.microsoft.com/.default"
GRAPH_BASE  = "https://graph.microsoft.com/v1.0"

app = func.FunctionApp()


# ── Timer trigger — every 15 minutes ─────────────────────────────────────────
@app.timer_trigger(
    schedule="0 */15 * * * *",
    arg_name="timer",
    run_on_startup=True,
)
def process_tender_schedule(timer: func.TimerRequest) -> None:
    if timer.past_due:
        logging.warning("Timer is past due — running now")
    logging.info("Tender schedule sync started at %s", datetime.now(timezone.utc).isoformat())
    token      = get_graph_token()
    xlsx_bytes = download_from_sharepoint(token)
    logging.info("Downloaded %d bytes from SharePoint", len(xlsx_bytes))
    result = process(xlsx_bytes)
    logging.info(
        "Complete — %d tenders written (%d EMEA, %d Americas, %d APAC)",
        result["total"],
        result["by_region"].get("EMEA", 0),
        result["by_region"].get("Americas", 0),
        result["by_region"].get("APAC", 0),
    )


# ── Graph API ─────────────────────────────────────────────────────────────────
def get_graph_token() -> str:
    cred  = ClientSecretCredential(TENANT_ID, CLIENT_ID, CLIENT_SECRET)
    return cred.get_token(GRAPH_SCOPE).token


def download_from_sharepoint(token: str) -> bytes:
    path = SHAREPOINT_FILE_PATH.replace(" ", "%20")
    url  = f"{GRAPH_BASE}/sites/{SHAREPOINT_SITE_ID}/drive/root:{path}:/content"
    res  = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    if res.status_code == 404:
        raise FileNotFoundError(f"File not found in SharePoint: {SHAREPOINT_FILE_PATH}")
    res.raise_for_status()
    return res.content


# ── Core processing ───────────────────────────────────────────────────────────
def process(xlsx_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)

    emea     = parse_emea(wb["EMEA Query"])
    americas = parse_us(wb["US Query"])
    apac     = parse_apac(wb["APAC Query"])
    wb.close()

    all_tenders = emea + americas + apac
    all_tenders = [t for t in all_tenders if t["title"]]

    by_region = {"EMEA": len(emea), "Americas": len(americas), "APAC": len(apac)}

    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total":        len(all_tenders),
        "by_region":    by_region,
        "tenders":      all_tenders,
        "parse_errors": [],
    }

    upload_json(output)
    return output


# ── EMEA Query parser ─────────────────────────────────────────────────────────
def parse_emea(ws) -> list[dict]:
    """
    Columns: CUSTOMER NUMBER, CUSTOMER NAME, CUSTOMER TYPE, ACCOUNT MANAGER,
    TENDER LEAD, TENDER NAME, TENDER STATUS (OPEN/CLOSED/AWARDED),
    TENDER_REMARKS (verbose stage), RFP VOLUME (USG), NO. OF ROUNDS,
    NO. OF LOCATIONS, Upcoming Due Date, CONTRACT START DATE, CONTRACT END DATE,
    ROUND 1 DEADLINE, ROUND 1 SUBMITTED … COMMENTS, Date Filter
    """
    rows_iter = ws.iter_rows(values_only=True)
    raw_h = next(rows_iter, None)
    if not raw_h:
        return []
    headers = [str(c).strip().upper() if c else "" for c in raw_h]

    def col(name):
        try: return headers.index(name)
        except ValueError: return None

    ic_num  = col("CUSTOMER NUMBER")
    ic_name = col("CUSTOMER NAME")
    ic_type = col("CUSTOMER TYPE")
    ic_mgr  = col("ACCOUNT MANAGER")
    ic_lead = col("TENDER LEAD")
    ic_tname= col("TENDER NAME")
    ic_stat = col("TENDER STATUS")
    ic_rem  = col("TENDER_REMARKS")
    ic_vol  = col("RFP VOLUME (USG)")
    ic_rds  = col("NO. OF ROUNDS")
    ic_locs = col("NO. OF LOCATIONS")
    ic_due  = col("UPCOMING DUE DATE")
    ic_cs   = col("CONTRACT START DATE")
    ic_ce   = col("CONTRACT END DATE")

    out = []
    for row in rows_iter:
        if not any(c for c in row):
            continue
        def g(i):
            return row[i] if i is not None and i < len(row) else None

        name = clean(g(ic_name))
        if not name:
            continue

        due = parse_date(g(ic_due))
        out.append({
            "id":             clean(g(ic_num)),
            "title":          name,
            "tender_name":    clean(g(ic_tname)),
            "status":         clean(g(ic_stat)),
            "remarks":        clean(g(ic_rem)),
            "assignee":       clean(g(ic_mgr)),
            "tender_lead":    clean(g(ic_lead)),
            "owner":          "",
            "deadline":       due,
            "days_remaining": days_rem(due),
            "contract_start": parse_date(g(ic_cs)),
            "contract_end":   parse_date(g(ic_ce)),
            "volume_usg":     to_float(g(ic_vol)),
            "locations":      to_int(g(ic_locs)),
            "rounds":         clean(g(ic_rds)),
            "customer_type":  clean(g(ic_type)),
            "region":         "EMEA",
        })
    return out


# ── US Query parser ───────────────────────────────────────────────────────────
def parse_us(ws) -> list[dict]:
    """
    Columns: Upcoming Due Date, Start Date, End Date, Completed Date,
    Request Date, Customer, Tender name, Cust Region, Location Region,
    Tender Status (verbose), Rounds, Round, Sales rep, Owner, Column1
    """
    rows_iter = ws.iter_rows(values_only=True)
    raw_h = next(rows_iter, None)
    if not raw_h:
        return []
    headers = [str(c).strip() if c else "" for c in raw_h]

    def col(name):
        try: return headers.index(name)
        except ValueError: return None

    ic_due   = col("Upcoming Due Date")
    ic_start = col("Start Date")
    ic_end   = col("End Date")
    ic_cust  = col("Customer")
    ic_tname = col("Tender name")
    ic_cregion= col("Cust Region")
    ic_lregion= col("Location Region")
    ic_stat  = col("Tender Status")
    ic_rounds= col("Rounds")
    ic_sales = col("Sales rep")
    ic_owner = col("Owner")

    out = []
    for row in rows_iter:
        if not any(c for c in row):
            continue
        def g(i):
            return row[i] if i is not None and i < len(row) else None

        name   = clean(g(ic_cust))
        status = clean(g(ic_stat))
        if not name or not status:
            continue

        # Derive simple status from verbose stage string
        sl = status.lower()
        if "awarded" in sl:
            simple = "AWARDED"
        elif "closed" in sl or "no bid" in sl or "no win" in sl:
            simple = "CLOSED"
        else:
            simple = "OPEN"

        due = parse_date(g(ic_due))
        out.append({
            "id":             "",
            "title":          name,
            "tender_name":    clean(g(ic_tname)),
            "status":         simple,
            "remarks":        status,
            "assignee":       clean(g(ic_sales)),
            "tender_lead":    "",
            "owner":          clean(g(ic_owner)),
            "deadline":       due,
            "days_remaining": days_rem(due),
            "contract_start": parse_date(g(ic_start)),
            "contract_end":   parse_date(g(ic_end)),
            "volume_usg":     None,
            "locations":      clean(g(ic_lregion)),
            "rounds":         clean(g(ic_rounds)),
            "customer_type":  "",
            "region":         "Americas",
        })
    return out


# ── APAC Query parser ─────────────────────────────────────────────────────────
def parse_apac(ws) -> list[dict]:
    """
    Columns: Upcoming Due Date, Start Date, End Date, Completed Date,
    Request Date, Customer, Tender name, Cust Region, Location Region,
    Tender Status, Awarded Locations, Margin, Sales rep, Owner
    """
    rows_iter = ws.iter_rows(values_only=True)
    raw_h = next(rows_iter, None)
    if not raw_h:
        return []
    headers = [str(c).strip() if c else "" for c in raw_h]

    def col(name):
        # APAC headers may have leading space
        for i, h in enumerate(headers):
            if h.strip() == name:
                return i
        return None

    ic_due    = col("Upcoming Due Date")
    ic_start  = col("Start Date")
    ic_end    = col("End Date")
    ic_cust   = col("Customer")
    ic_tname  = col("Tender name")
    ic_lregion= col("Location Region")
    ic_stat   = col("Tender Status")
    ic_sales  = col("Sales rep")
    ic_owner  = col("Owner")
    ic_margin = col("Margin")

    out = []
    for row in rows_iter:
        if not any(c for c in row):
            continue
        def g(i):
            return row[i] if i is not None and i < len(row) else None

        name   = clean(g(ic_cust))
        status = clean(g(ic_stat))
        if not name or not status:
            continue

        sl = status.lower()
        if "awarded" in sl:
            simple = "AWARDED"
        elif "closed" in sl or "no bid" in sl or "no win" in sl or "no support" in sl:
            simple = "CLOSED"
        else:
            simple = "OPEN"

        due = parse_date(g(ic_due))
        out.append({
            "id":             "",
            "title":          name,
            "tender_name":    clean(g(ic_tname)),
            "status":         simple,
            "remarks":        status,
            "assignee":       clean(g(ic_sales)),
            "tender_lead":    "",
            "owner":          clean(g(ic_owner)),
            "deadline":       due,
            "days_remaining": days_rem(due),
            "contract_start": parse_date(g(ic_start)),
            "contract_end":   parse_date(g(ic_end)),
            "volume_usg":     None,
            "locations":      clean(g(ic_lregion)),
            "rounds":         None,
            "customer_type":  "",
            "region":         "APAC",
            "margin":         clean(g(ic_margin)),
        })
    return out


# ── Helpers ───────────────────────────────────────────────────────────────────
def clean(v, fallback: str = "") -> str:
    if v is None:
        return fallback
    s = str(v).strip()
    return s if s and s != "None" else fallback


def parse_date(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    # Excel serial number (fallback for read_only mode edge cases)
    try:
        n = int(float(str(v)))
        return (datetime(1899, 12, 30) + timedelta(days=n)).date().isoformat()
    except Exception:
        return None


def days_rem(d_str: str | None) -> int | None:
    if not d_str:
        return None
    try:
        return (date.fromisoformat(d_str) - date.today()).days
    except Exception:
        return None


def to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def to_int(v) -> int | None:
    f = to_float(v)
    return int(f) if f is not None else None


def upload_json(data: dict) -> None:
    blob_service = BlobServiceClient.from_connection_string(STORAGE_CONNECTION)
    container    = blob_service.get_container_client(OUTPUT_CONTAINER)
    try:
        container.create_container(public_access="blob")
    except Exception:
        pass

    payload = json.dumps(data, indent=2, default=str, ensure_ascii=False).encode("utf-8")
    blob_service.get_blob_client(OUTPUT_CONTAINER, OUTPUT_BLOB_NAME).upload_blob(
        payload,
        overwrite=True,
        content_settings=ContentSettings(
            content_type="application/json",
            cache_control="no-cache, must-revalidate",
        ),
    )
    logging.info("Uploaded %d bytes → %s/%s", len(payload), OUTPUT_CONTAINER, OUTPUT_BLOB_NAME)
